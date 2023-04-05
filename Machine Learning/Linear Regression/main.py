from multi_regression import Multiple_Linear_Regression
from simple_regression import Simple_Linear_Regression
import numpy as np
from openpyxl import load_workbook


def main():

    total_iterations = 10
    learning_rate = 6e-7
    lambda_ = 2

    # ------------------------------------------TEST--------------------------------------------------

    # One variable
    # X_train = np.array([1.0, 1.7, 2.0, 2.5, 3.0, 3.2])
    # Y_train = np.array([250, 300, 480,  430,   630, 730])

    # srg = Simple_Linear_Regression(X_train, Y_train, learning_rate, total_iterations, lambda_)
    # w, b = srg.show_parameters()
    # srg.predict(X_train, Y_train, w, b)

    # Multi variable
    # X_train = np.array([[2104, 5, 1, 45], [1416, 3, 2, 40], [852, 2, 1, 35]])
    # Y_train = np.array([460, 232, 178])

    # rgm = Multiple_Linear_Regression(X_train, Y_train, learning_rate,
    #                        total_iterations, lambda_)
    # w, b = rgm.show_parameters()
    # rgm.predict(X_train, Y_train, w, b)

    # ------------------------------------------TEST-END----------------------------------------------

    workbook = load_workbook("dataset.xlsx")
    sheet = workbook.active

    row_count = sheet.max_row

    x_train = np.zeros(row_count)
    y_train = np.zeros(row_count)

    for i in range(row_count):
        x_train[i] = sheet.cell(row=i+1, column=1).value
        y_train[i] = sheet.cell(row=i+1, column=2).value

    srg = Simple_Linear_Regression(
        x_train, y_train, learning_rate, total_iterations, lambda_)
    w, b = srg.show_parameters()
    srg.predict(x_train, y_train, w, b)


if __name__ == "__main__":
    main()
